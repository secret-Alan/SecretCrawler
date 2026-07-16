"""数据管道（crawler.pipeline）单元测试。

覆盖 spec 中「数据管道与存储」需求：CSV / JSON / Excel / SQLite / NONE，
以及 MultiPipeline 在 FileDownload 失败时的隔离性。全部离线，使用 tmp_path。
"""
from __future__ import annotations

import csv
import json
import sqlite3

import pytest

from crawler.models import CrawlConfig, ResultItem
from crawler.pipeline import PipelineFactory


def _item(url: str, **fields) -> ResultItem:
    """构造一条 ResultItem。"""
    return ResultItem(url=url, fields=dict(fields))


def test_csv_pipeline_writes_bom_and_rows(tmp_path):
    """CSV：UTF-8 BOM + 表头 + 数据行。"""
    config = CrawlConfig(storage_format="CSV", output_path=str(tmp_path))
    pipe = PipelineFactory.build(config)
    pipe.process(_item("https://e.com/1", title="A", author="Z"))
    pipe.process(_item("https://e.com/2", title="B", author="Y"))
    pipe.close()

    csv_files = list(tmp_path.glob("*.csv"))
    assert csv_files, "应生成 CSV 文件"
    data = csv_files[0].read_bytes()
    # BOM 存在（utf-8-sig 写入）
    assert data.startswith(b"\xef\xbb\xbf")

    # 行内容正确（utf-8-sig 解码会剥离 BOM）
    reader = csv.reader(data.decode("utf-8-sig").splitlines())
    rows = list(reader)
    assert rows[0] == ["title", "author"]
    assert rows[1] == ["A", "Z"]
    assert rows[2] == ["B", "Y"]


def test_json_pipeline_writes_array(tmp_path):
    """JSON：数组格式，字段值正确。"""
    out = tmp_path / "out.json"
    config = CrawlConfig(storage_format="JSON", output_path=str(out))
    pipe = PipelineFactory.build(config)
    pipe.process(_item("https://e.com/1", title="A"))
    pipe.process(_item("https://e.com/2", title="B"))
    pipe.close()

    payload = json.loads(out.read_text(encoding="utf-8"))
    assert isinstance(payload, list)
    assert len(payload) == 2
    assert payload[0]["url"] == "https://e.com/1"
    assert payload[0]["fields"]["title"] == "A"
    assert payload[1]["url"] == "https://e.com/2"
    assert payload[1]["fields"]["title"] == "B"


def test_excel_pipeline_writes_sheet(tmp_path):
    """Excel：sheet 名 "Results"，含表头 + 数据行。"""
    pytest.importorskip("openpyxl")
    from openpyxl import load_workbook

    out = tmp_path / "out.xlsx"
    config = CrawlConfig(storage_format="EXCEL", output_path=str(out))
    pipe = PipelineFactory.build(config)
    pipe.process(_item("https://e.com/1", title="A"))
    pipe.process(_item("https://e.com/2", title="B"))
    pipe.close()

    wb = load_workbook(out)
    assert "Results" in wb.sheetnames
    ws = wb["Results"]
    # 第 1 行表头
    assert [c.value for c in ws[1]] == ["title"]
    # 数据行
    assert [c.value for c in ws[2]] == ["A"]
    assert [c.value for c in ws[3]] == ["B"]


def test_sqlite_pipeline_dynamic_columns(tmp_path):
    """SQLite：crawl_results 表，动态添加新列。"""
    out = tmp_path / "out.db"
    config = CrawlConfig(storage_format="SQLITE", output_path=str(out))
    pipe = PipelineFactory.build(config)
    pipe.process(_item("https://e.com/1", title="A"))
    # 第二条结果带新字段 author → 应 ALTER TABLE 添加列
    pipe.process(_item("https://e.com/2", title="B", author="X"))
    pipe.close()

    with sqlite3.connect(out) as conn:
        cols = [r[1] for r in conn.execute("PRAGMA table_info(crawl_results)").fetchall()]
        assert "id" in cols
        assert "url" in cols
        assert "title" in cols
        # 动态添加的列
        assert "author" in cols

        rows = conn.execute(
            "SELECT url, title, author FROM crawl_results ORDER BY id"
        ).fetchall()

    assert len(rows) == 2
    # 第一行在新列添加前插入，author 为 NULL
    assert rows[0][0] == "https://e.com/1"
    assert rows[0][1] == "A"
    assert rows[0][2] is None
    # 第二行带 author 值
    assert rows[1] == ("https://e.com/2", "B", "X")


def test_none_pipeline_creates_no_file(tmp_path):
    """NONE：不创建文件，process 不抛异常。"""
    config = CrawlConfig(storage_format="NONE", output_path=str(tmp_path))
    pipe = PipelineFactory.build(config)
    pipe.process(_item("https://e.com/1", title="A"))
    pipe.close()
    # 目录内无任何文件
    assert not any(tmp_path.iterdir())


def test_download_files_isolated_from_storage(tmp_path, monkeypatch):
    """download_files=True 时，FileDownload 失败不影响存储管道写入。"""
    pytest.importorskip("requests")
    import requests as _requests

    # 让 requests.get 立即失败，保持离线
    def _boom(*args, **kwargs):
        raise _requests.ConnectionError("offline test")

    monkeypatch.setattr(_requests, "get", _boom)

    out = tmp_path / "out.csv"
    config = CrawlConfig(
        storage_format="CSV",
        output_path=str(out),
        download_files=True,
        download_url_regex=r"https?://[^\s\"'<>]+",
        download_dir=str(tmp_path / "downloads"),
        timeout=1.0,
    )
    pipe = PipelineFactory.build(config)
    # item.url 匹配下载正则，但 requests.get 被替换为抛异常 → 下载失败被隔离
    pipe.process(_item("http://127.0.0.1:1/unreachable", title="A"))
    pipe.close()

    # 存储管道（CSV）应仍写出文件
    assert out.exists()
    reader = csv.reader(out.read_text(encoding="utf-8-sig").splitlines())
    rows = list(reader)
    assert rows[0] == ["title"]
    assert rows[1] == ["A"]
